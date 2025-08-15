import openpyxl
import sys
import os

def check_excel_content(file_path):
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"Excel 文件: {file_path}")
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n工作表: {sheet_name}")
            print(f"最大行数: {sheet.max_row}")
            print(f"最大列数: {sheet.max_column}")
            
            # 显示前6行内容
            for row in range(1, min(7, sheet.max_row + 1)):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "NULL")
                print(f"  行 {row}: {' | '.join(row_data)}")
                
    except Exception as e:
        print(f"读取 Excel 文件时发生错误: {e}")

if __name__ == "__main__":
    excel_path = "Assets/ExcelConfigs/Example.xlsx"
    check_excel_content(excel_path)
